from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties


SOURCE = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_OWNER_WORKLOAD_FIXED_source.xlsx")
OUTPUT = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_LIVE_FORMULAS.xlsx")

REPLACEMENTS = {
    "CADTracker_Tracker_ID": "CADTracker[Tracker ID]",
    "CADTracker_Batch_ID": "CADTracker[Batch ID]",
    "CADTracker_WE": "CADTracker[WE]",
    "CADTracker_Project": "CADTracker[Project / Collection]",
    "CADTracker_Item": "CADTracker[Item / Design No.]",
    "CADTracker_Request_Type": "CADTracker[Request Type]",
    "CADTracker_Deliverable_Type": "CADTracker[Deliverable Type]",
    "CADTracker_Task": "CADTracker[Task]",
    "CADTracker_Qty": "CADTracker[Qty]",
    "CADTracker_Status": "CADTracker[Status]",
    "CADTracker_Owner": "CADTracker[Owner]",
    "CADTracker_Priority": "CADTracker[Priority]",
    "CADTracker_Start_Date": "CADTracker[Start Date]",
    "CADTracker_Plan_End": "CADTracker[Plan End]",
}


def main():
    wb = load_workbook(SOURCE)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    ws = wb["Dashboard"]

    changed = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                updated = formula
                for old, new in REPLACEMENTS.items():
                    updated = updated.replace(old, new)
                if updated != formula:
                    cell.value = updated
                    changed += 1

    if changed == 0:
        raise RuntimeError("No Dashboard formulas were updated.")

    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    dashboard = check["Dashboard"]
    errors = []
    old_refs = []
    for row in dashboard.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for token in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
                    if token in cell.value:
                        errors.append((cell.coordinate, cell.value))
                for old in REPLACEMENTS:
                    if old in cell.value:
                        old_refs.append((cell.coordinate, cell.value))

    if errors:
        raise RuntimeError(f"Formula error tokens found: {errors[:10]}")
    if old_refs:
        raise RuntimeError(f"Old named refs still found: {old_refs[:10]}")

    print(f"{OUTPUT} updated_formulas={changed}")


if __name__ == "__main__":
    main()
