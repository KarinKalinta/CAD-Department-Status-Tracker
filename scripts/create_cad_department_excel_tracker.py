from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "CAD_Department_Plan_of_Action_auto_timeline_no_hash_dates.xlsx"

INPUT_END_ROW = 60
TIMELINE_TITLE_ROW = 64
TIMELINE_HEADER_ROW = 65
TIMELINE_START_ROW = 66

INPUT_HEADERS = [
    "WE",
    "Project / Collection",
    "Request Type",
    "Deliverable Type",
    "Task",
    "Status",
    "Owner",
    "Start Date",
    "Plan End",
    "Priority",
    "Estimated Time",
    "Notes",
    "Link / File",
]

TIMELINE_HEADERS = ["Project / Collection", "Task", "Owner", "Status"]

LISTS = {
    "Request Type": ["Initial Request", "New Request", "Revision", "Follow-up"],
    "Project / Collection": [
        "C0553-F",
        "C0731 Mireyas",
        "C0634-X",
        "C0634",
        "Organic",
        "Men's",
        "Cluster",
        "RhinoArtisan",
        "PD CAD Investigation",
        "Cosmic Collection",
        "Small Scale Design",
        "Custom Order",
        "Internal CAD",
        "Other",
    ],
    "Task": [
        "Sketch",
        "Existing sketch",
        "3D set",
        "Photo",
        "CAD",
        "CAD Render Files",
        "CAD PD Files",
        "Presentation",
        "PD Update Progress",
        "Render",
        "Sample",
        "R&D",
        "Finding Data",
        "Update Production to CCU and OPC",
        "Test with real job",
    ],
    "Deliverable Type": [
        "Sketch",
        "Existing Sketch",
        "3D Asset",
        "Photo",
        "CAD File",
        "Email",
        "Render",
        "Sample",
        "Document",
        "Presentation",
        "STL / Print File",
        "Other",
    ],
    "Estimated Time": [
        "15 minutes",
        "20 minutes",
        "30 to 45 minutes",
        "1 hour",
        "2 to 3 hours",
        "Half day",
        "1 day",
        "2 days",
        "This week",
    ],
    "Priority": ["High", "Medium", "Low"],
    "Status": ["In Progress", "Waiting", "Done", "Blocked", "Cancelled"],
    "Owner": ["CAD", "Boom", "Production", "Designer", "Manager", "Customer", "Other"],
}

EXAMPLES = [
    [
        "2026-W27",
        "C0553-F",
        "Initial Request",
        "CAD File",
        "CAD",
        "In Progress",
        "CAD",
        date(2026, 6, 29),
        date(2026, 7, 2),
        "High",
        "2 to 3 hours",
        "Active CAD task.",
        "",
    ],
    [
        "2026-W27",
        "RhinoArtisan",
        "Follow-up",
        "Document",
        "Test with real job",
        "Waiting",
        "Boom",
        date(2026, 7, 1),
        date(2026, 7, 4),
        "Medium",
        "Half day",
        "Waiting for review input.",
        "",
    ],
    [
        "2026-W27",
        "Cosmic Collection",
        "Initial Request",
        "Presentation",
        "Presentation",
        "Done",
        "Designer",
        date(2026, 7, 6),
        date(2026, 7, 7),
        "Low",
        "1 day",
        "Completed presentation task.",
        "",
    ],
    [
        "2026-W27",
        "Custom Order",
        "Revision",
        "Email",
        "Update Production to CCU and OPC",
        "Blocked",
        "Production",
        date(2026, 7, 8),
        date(2026, 7, 10),
        "High",
        "30 to 45 minutes",
        "Blocked until production confirms.",
        "",
    ],
    [
        "2026-W28",
        "C0731 Mireyas",
        "New Request",
        "Sketch",
        "Sketch",
        "Cancelled",
        "Boom",
        date(2026, 7, 13),
        date(2026, 7, 15),
        "Medium",
        "2 to 3 hours",
        "Cancelled sample row.",
        "",
    ],
]


def workdays_from_current_week(start_date: date, weeks: int = 3) -> list[date]:
    monday = start_date - timedelta(days=start_date.weekday())
    days: list[date] = []
    cursor = monday
    while len(days) < weeks * 5:
        if cursor.weekday() < 5:
            days.append(cursor)
        cursor += timedelta(days=1)
    return days


def week_ranges_from_current_week(start_date: date, weeks: int = 3) -> list[tuple[date, date, str]]:
    monday = start_date - timedelta(days=start_date.weekday())
    ranges: list[tuple[date, date, str]] = []
    for offset in range(weeks):
        week_start = monday + timedelta(days=offset * 7)
        week_end = week_start + timedelta(days=4)
        label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}"
        ranges.append((week_start, week_end, label))
    return ranges


def make_styles():
    thin = Side(style="thin", color="D9E2F3")
    return {
        "header_fill": PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78"),
        "title_fill": PatternFill(fill_type="solid", start_color="111827", end_color="111827"),
        "header_font": Font(name="Aptos", size=10, bold=True, color="FFFFFF"),
        "title_font": Font(name="Aptos", size=14, bold=True, color="FFFFFF"),
        "body_font": Font(name="Aptos", size=10),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def setup_input_section(ws, include_examples: bool):
    styles = make_styles()

    ws.append(INPUT_HEADERS)
    if include_examples:
        for row in EXAMPLES:
            ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{INPUT_END_ROW}"

    for cell in ws[1]:
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["border"]

    widths = [15, 22, 28, 26, 38, 18, 18, 18, 18, 16, 22, 56, 34]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width

    for row in ws.iter_rows(min_row=2, max_row=INPUT_END_ROW, min_col=1, max_col=len(INPUT_HEADERS)):
        for cell in row:
            cell.font = styles["body_font"]
            cell.alignment = Alignment(vertical="center", wrap_text=cell.column in (12, 13))
            cell.border = styles["border"]

    for row in range(2, INPUT_END_ROW + 1):
        ws[f"H{row}"].number_format = "[$-en-US]d mmm yy"
        ws[f"I{row}"].number_format = "[$-en-US]d mmm yy"

    add_dropdowns(ws)
    add_input_conditional_formatting(ws)
    color_existing_input_cells(ws, max_row=INPUT_END_ROW)


def setup_timeline_section(ws, timeline_weeks: list[tuple[date, date, str]]):
    styles = make_styles()
    timeline_date_start_col = len(TIMELINE_HEADERS) + 1
    timeline_end_col = len(TIMELINE_HEADERS) + len(timeline_weeks)
    timeline_end_row = TIMELINE_START_ROW + (INPUT_END_ROW - 2)

    title_range = f"A{TIMELINE_TITLE_ROW}:{get_column_letter(timeline_end_col)}{TIMELINE_TITLE_ROW}"
    ws.merge_cells(title_range)
    title_cell = ws[f"A{TIMELINE_TITLE_ROW}"]
    title_cell.value = "TIMELINE VIEW"
    title_cell.fill = styles["title_fill"]
    title_cell.font = styles["title_font"]
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[TIMELINE_TITLE_ROW].height = 24

    for column_index, header in enumerate(TIMELINE_HEADERS, 1):
        cell = ws.cell(row=TIMELINE_HEADER_ROW, column=column_index, value=header)
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["border"]

    for offset, (_week_start, _week_end, _label) in enumerate(timeline_weeks):
        column_index = timeline_date_start_col + offset
        if offset == 0:
            cell = ws.cell(
                row=TIMELINE_HEADER_ROW,
                column=column_index,
                value=(
                    '=IF(COUNT($H$2:$H$60)>0,'
                    'MIN($H$2:$H$60)-WEEKDAY(MIN($H$2:$H$60),2)+1,'
                    'TODAY()-WEEKDAY(TODAY(),2)+1)'
                ),
            )
        else:
            previous_column = get_column_letter(column_index - 1)
            cell = ws.cell(
                row=TIMELINE_HEADER_ROW,
                column=column_index,
                value=f"={previous_column}${TIMELINE_HEADER_ROW}+7",
            )
        cell.fill = styles["header_fill"]
        cell.font = styles["header_font"]
        cell.number_format = "[$-en-US]d mmm"
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["border"]
        ws.column_dimensions[get_column_letter(column_index)].width = max(
            ws.column_dimensions[get_column_letter(column_index)].width or 0,
            18,
        )

    ws.row_dimensions[TIMELINE_HEADER_ROW].height = 42
    for column_index, width in enumerate([28, 38, 18, 18], 1):
        ws.column_dimensions[get_column_letter(column_index)].width = max(
            ws.column_dimensions[get_column_letter(column_index)].width or 0,
            width,
        )

    for row_index, source_row in enumerate(range(2, INPUT_END_ROW + 1), TIMELINE_START_ROW):
        formulas = [
            f'=IF($B{source_row}="","",$B{source_row})',
            f'=IF($E{source_row}="","",$E{source_row})',
            f'=IF($G{source_row}="","",$G{source_row})',
            f'=IF($F{source_row}="","",$F{source_row})',
        ]
        for column_index, formula in enumerate(formulas, 1):
            cell = ws.cell(row=row_index, column=column_index, value=formula)
            cell.font = styles["body_font"]
            cell.alignment = Alignment(vertical="center", wrap_text=column_index == 2)
            cell.border = styles["border"]

        for column_index in range(timeline_date_start_col, timeline_end_col + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.border = styles["border"]

    add_timeline_formatting(ws, timeline_weeks, timeline_end_row)


def add_dropdowns(ws):
    dropdown_columns = {
        "B": "B",
        "C": "A",
        "D": "D",
        "E": "C",
        "F": "G",
        "G": "H",
        "J": "F",
        "K": "E",
    }

    for target_col, source_col in dropdown_columns.items():
        formula = f"='Lists hidden'!${source_col}$2:${source_col}$100"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Choose a value from the dropdown list."
        validation.errorTitle = "Invalid value"
        ws.add_data_validation(validation)
        validation.add(f"{target_col}2:{target_col}{INPUT_END_ROW}")


def add_timeline_formatting(ws, timeline_weeks: list[tuple[date, date, str]], timeline_end_row: int):
    timeline_start_col = len(TIMELINE_HEADERS) + 1

    status_colors = {
        "In Progress": "3B82F6",
        "Waiting": "FDE68A",
        "Done": "86EFAC",
        "Blocked": "FCA5A5",
        "Cancelled": "D1D5DB",
    }

    for offset, (_week_start, _week_end, _label) in enumerate(timeline_weeks):
        column_letter = get_column_letter(timeline_start_col + offset)
        week_range = f"{column_letter}{TIMELINE_START_ROW}:{column_letter}{timeline_end_row}"

        for status, fill_color in status_colors.items():
            ws.conditional_formatting.add(
                week_range,
                FormulaRule(
                    formula=[
                        (
                            f'=AND({column_letter}${TIMELINE_HEADER_ROW}<=$I2,'
                            f'{column_letter}${TIMELINE_HEADER_ROW}+4>=$H2,'
                            f'$H2<>"",$I2<>"",$F2="{status}")'
                        )
                    ],
                    fill=PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color),
                ),
            )


def color_existing_input_cells(ws, max_row: int):
    fills = {
        "High": ("FFEDD5", "9A3412"),
        "Medium": ("FEF3C7", "92400E"),
        "Low": ("E5E7EB", "374151"),
        "In Progress": ("DBEAFE", "1E3A8A"),
        "Waiting": ("FEF3C7", "92400E"),
        "Done": ("DCFCE7", "166534"),
        "Blocked": ("FEE2E2", "991B1B"),
        "Cancelled": ("E5E7EB", "374151"),
        "CAD": ("FED7AA", "9A3412"),
        "Boom": ("B91C1C", "FFFFFF"),
        "Production": ("BAE6FD", "075985"),
        "Designer": ("E9D5FF", "6B21A8"),
    }

    for row in range(2, max_row + 1):
        for column in ("F", "G", "J"):
            cell = ws[f"{column}{row}"]
            if cell.value in fills:
                fill_color, font_color = fills[cell.value]
                cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
                cell.font = Font(name="Aptos", size=10, color=font_color)


def add_input_conditional_formatting(ws):
    rules = {
        "F": {
            "In Progress": ("DBEAFE", "1E3A8A"),
            "Waiting": ("FEF3C7", "92400E"),
            "Done": ("DCFCE7", "166534"),
            "Blocked": ("FEE2E2", "991B1B"),
            "Cancelled": ("E5E7EB", "374151"),
        },
        "G": {
            "CAD": ("FED7AA", "9A3412"),
            "Boom": ("B91C1C", "FFFFFF"),
            "Production": ("BAE6FD", "075985"),
            "Designer": ("E9D5FF", "6B21A8"),
            "Manager": ("DCFCE7", "166534"),
            "Customer": ("DBEAFE", "1E3A8A"),
            "Other": ("E5E7EB", "374151"),
        },
        "J": {
            "High": ("FFEDD5", "9A3412"),
            "Medium": ("FEF3C7", "92400E"),
            "Low": ("E5E7EB", "374151"),
        },
    }

    for column, values in rules.items():
        target_range = f"{column}2:{column}{INPUT_END_ROW}"
        for value, (fill_color, font_color) in values.items():
            ws.conditional_formatting.add(
                target_range,
                FormulaRule(
                    formula=[f'=${column}2="{value}"'],
                    fill=PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color),
                    font=Font(name="Aptos", size=10, color=font_color),
                ),
            )


def setup_lists_sheet(ws):
    styles = make_styles()
    for column_index, (header, values) in enumerate(LISTS.items(), 1):
        header_cell = ws.cell(row=1, column=column_index, value=header)
        header_cell.fill = styles["header_fill"]
        header_cell.font = styles["header_font"]
        header_cell.alignment = Alignment(horizontal="center")
        for row_index, value in enumerate(values, 2):
            ws.cell(row=row_index, column=column_index, value=value)
        ws.column_dimensions[get_column_letter(column_index)].width = 26
    ws.sheet_state = "hidden"


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)

    timeline_weeks = week_ranges_from_current_week(date(2026, 6, 30), weeks=8)

    plan = wb.create_sheet("Plan of Action")
    examples = wb.create_sheet("CAD Examples")
    lists_ws = wb.create_sheet("Lists hidden")

    setup_input_section(plan, include_examples=False)
    setup_timeline_section(plan, timeline_weeks)

    setup_input_section(examples, include_examples=True)
    setup_timeline_section(examples, timeline_weeks)

    setup_lists_sheet(lists_ws)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    plan.sheet_view.topLeftCell = "A1"
    plan.sheet_view.selection[0].activeCell = "A2"
    plan.sheet_view.selection[0].sqref = "A2"
    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    create_workbook()
