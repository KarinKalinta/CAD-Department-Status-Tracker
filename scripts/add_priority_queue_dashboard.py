from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties


SOURCE = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_FIXED_source.xlsx")
OUTPUT = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_FINAL.xlsx")


def tcol(name: str) -> str:
    return f"CADTracker[{name}]"


def style_header(cell, fill="1F4E78"):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_range(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            if cell.row > min_row:
                cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def queue_index_formula(row_number: int) -> str:
    task = tcol("Task")
    priority = tcol("Priority")
    status = tcol("Status")
    plan_end = tcol("Plan End")
    n = row_number - 10
    criteria = (
        f'({task}<>"")*'
        f'(({priority}="Urgent")+({priority}="High")+({status}="Blocked")+'
        f'(({plan_end}<TODAY())*({status}<>"Done")*({status}<>"Cancelled"))+'
        f'(({plan_end}>=TODAY())*({plan_end}<=TODAY()+7)*({status}<>"Done")*({status}<>"Cancelled"))>0)'
    )
    return f'=IFERROR(AGGREGATE(15,6,(ROW({task})-MIN(ROW({task}))+1)/({criteria}),{n}),"")'


def queue_value_formula(index_cell: str, column_name: str, default_qty=False) -> str:
    value = f"INDEX({tcol(column_name)},${index_cell})"
    if default_qty:
        return f'=IF(${index_cell}="","",IF({value}="",1,{value}))'
    return f'=IF(${index_cell}="","",{value})'


def find_cell(ws, text):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell
    return None


def add_priority_queue(ws):
    insert_at = 9
    rows_to_insert = 18
    ws.insert_rows(insert_at, rows_to_insert)

    title_row = 9
    header_row = 10
    first_data = 11
    last_data = 25

    ws.cell(title_row, 1, "Priority Queue")
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=8)
    style_header(ws.cell(title_row, 1), "111827")
    ws.cell(title_row, 1).font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")

    headers = [
        "Priority",
        "Project / Collection",
        "Item / Design No.",
        "Task",
        "Qty",
        "Status",
        "Owner",
        "Plan End",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(header_row, col, header)
        style_header(ws.cell(header_row, col), "1F4E78")

    for row in range(first_data, last_data + 1):
        helper_cell = f"I{row}"
        ws.cell(row, 9, queue_index_formula(row))
        mappings = [
            (1, "Priority", False),
            (2, "Project / Collection", False),
            (3, "Item / Design No.", False),
            (4, "Task", False),
            (5, "Qty", True),
            (6, "Status", False),
            (7, "Owner", False),
            (8, "Plan End", False),
        ]
        for col, source_col, is_qty in mappings:
            ws.cell(row, col, queue_value_formula(helper_cell, source_col, default_qty=is_qty))
        ws.cell(row, 8).number_format = "[$-en-US]d mmm yy"

    ws.column_dimensions["I"].hidden = True
    style_range(ws, title_row, last_data, 1, 9)


def fix_batch_count(ws):
    cell = find_cell(ws, "Batch Count")
    if not cell:
        return
    value_cell = ws.cell(cell.row + 1, cell.column)
    batch = tcol("Batch ID")
    task = tcol("Task")
    value_cell.value = f'=SUMPRODUCT(({batch}<>"")*({task}<>"")/COUNTIFS({batch},{batch}&"",{task},"<>"))'


def fix_batch_summary(ws):
    title = find_cell(ws, "Batch Summary")
    if not title:
        return

    start_row = title.row
    header_row = start_row + 1
    first_data = start_row + 2
    last_data = first_data + 19
    start_col = title.column
    batch_col_letter = get_column_letter(start_col)

    expected_headers = [
        "Batch ID",
        "Project / Collection",
        "Total Items",
        "Done Items",
        "Blocked Items",
        "Progress %",
        "Overall Status",
    ]
    for offset, header in enumerate(expected_headers):
        ws.cell(header_row, start_col + offset, header)
        style_header(ws.cell(header_row, start_col + offset), "1F4E78")

    batch = tcol("Batch ID")
    project = tcol("Project / Collection")
    task = tcol("Task")
    qty = tcol("Qty")
    status = tcol("Status")

    for row in range(first_data, last_data + 1):
        batch_cell = ws.cell(row, start_col).coordinate
        previous_range = f"${batch_col_letter}${first_data - 1}:{batch_col_letter}{row - 1}"
        total_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>")+COUNTIFS({batch},{batch_cell},{task},"<>",{qty},"")'
        done_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>",{status},"Done")+COUNTIFS({batch},{batch_cell},{task},"<>",{status},"Done",{qty},"")'
        blocked_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>",{status},"Blocked")+COUNTIFS({batch},{batch_cell},{task},"<>",{status},"Blocked",{qty},"")'

        ws.cell(row, start_col, f'=IFERROR(INDEX({batch},MATCH(0,COUNTIF({previous_range},{batch})+({batch}="")+({task}=""),0)),"")')
        ws.cell(row, start_col + 1, f'=IF({batch_cell}="","",INDEX({project},MATCH({batch_cell},{batch},0)))')
        ws.cell(row, start_col + 2, f'=IF({batch_cell}="","",{total_items})')
        ws.cell(row, start_col + 3, f'=IF({batch_cell}="","",{done_items})')
        ws.cell(row, start_col + 4, f'=IF({batch_cell}="","",{blocked_items})')
        ws.cell(row, start_col + 5, f'=IFERROR(({done_items})/({total_items}),"")')
        ws.cell(row, start_col + 6, (
            f'=IF({batch_cell}="","",IF(({done_items})=({total_items}),"Done",'
            f'IF(({blocked_items})>0,"Blocked",'
            f'IF(COUNTIFS({batch},{batch_cell},{status},"In Progress")+COUNTIFS({batch},{batch_cell},{status},"Checking")>0,"In Progress",'
            f'IF(COUNTIFS({batch},{batch_cell},{status},"Waiting")=COUNTIFS({batch},{batch_cell},{task},"<>"),"Waiting","Mixed")))))'
        ))
        ws.cell(row, start_col + 5).number_format = "0%"

    style_range(ws, start_row, last_data, start_col, start_col + 6)


def main():
    wb = load_workbook(SOURCE)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    ws = wb["Dashboard"]

    add_priority_queue(ws)
    fix_batch_count(ws)
    fix_batch_summary(ws)

    ws.sheet_view.showGridLines = False
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    errors = []
    for row in check["Dashboard"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for token in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
                    if token in cell.value:
                        errors.append((cell.coordinate, cell.value))
    if errors:
        raise RuntimeError(errors[:20])
    print(OUTPUT)


if __name__ == "__main__":
    main()
