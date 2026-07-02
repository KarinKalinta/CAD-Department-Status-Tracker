from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties


SOURCE = Path("CAD_Status_Tracker_Batch_Aware_OWNERS_UPDATED_source.xlsx")
OUTPUT = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW.xlsx")

OWNERS = ["CAD1", "CAD2", "CAD3", "CAD4", "CAD5", "CAD6", "Outsource CAD", "Sdesign", "Design 2"]


def tcol(name: str) -> str:
    return f"CADTracker[{name}]"


def qty_active_formula(owner_cell=None):
    task = tcol("Task")
    qty = tcol("Qty")
    status = tcol("Status")
    owner = tcol("Owner")
    if owner_cell:
        return (
            f'=SUMIFS({qty},{task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled")'
            f'+COUNTIFS({task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled",{qty},"")'
        )
    return (
        f'=SUMIFS({qty},{task},"<>",{status},"<>Done",{status},"<>Cancelled")'
        f'+COUNTIFS({task},"<>",{status},"<>Done",{status},"<>Cancelled",{qty},"")'
    )


def queue_index_formula(row_number: int) -> str:
    task = tcol("Task")
    priority = tcol("Priority")
    status = tcol("Status")
    plan_end = tcol("Plan End")
    n = row_number - 8
    criteria = (
        f'({task}<>"")*'
        f'(({priority}="Urgent")+({priority}="High")+({status}="Blocked")+'
        f'(({plan_end}<TODAY())*({status}<>"Done")*({status}<>"Cancelled"))+'
        f'(({plan_end}>=TODAY())*({plan_end}<=TODAY()+7)*({status}<>"Done")*({status}<>"Cancelled"))>0)'
    )
    return f'=IFERROR(AGGREGATE(15,6,(ROW({task})-MIN(ROW({task}))+1)/({criteria}),{n}),"")'


def queue_value_formula(index_cell: str, column_name: str) -> str:
    return f'=IF(${index_cell}="","",INDEX({tcol(column_name)},${index_cell}))'


def batch_id_formula(row: int, first_row: int) -> str:
    previous_range = f"$A${first_row - 1}:A{row - 1}"
    batch = tcol("Batch ID")
    return f'=IFERROR(INDEX({batch},MATCH(0,COUNTIF({previous_range},{batch})+({batch}=""),0)),"")'


def set_header(cell, fill="1F4E78"):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_title(ws, row: int, text: str, last_col: int):
    ws.cell(row, 1, text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    set_header(ws.cell(row, 1), "111827")
    ws.cell(row, 1).font = Font(name="Aptos", size=14, bold=True, color="FFFFFF")


def style_block(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.font = cell.font.copy(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def rebuild_dashboard(ws):
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False

    task = tcol("Task")
    status = tcol("Status")
    priority = tcol("Priority")
    plan_end = tcol("Plan End")
    qty = tcol("Qty")
    owner = tcol("Owner")
    batch = tcol("Batch ID")
    project = tcol("Project / Collection")

    set_title(ws, 1, "CAD Batch Status Dashboard", 12)

    kpis = [
        ("Active Tasks", f'=COUNTIFS({task},"<>",{status},"<>Done",{status},"<>Cancelled")'),
        ("Active Items", qty_active_formula()),
        ("Urgent", f'=COUNTIFS({task},"<>",{priority},"Urgent",{status},"<>Done",{status},"<>Cancelled")'),
        ("Overdue", f'=COUNTIFS({task},"<>",{plan_end},"<"&TODAY(),{status},"<>Done",{status},"<>Cancelled")'),
        ("Blocked", f'=COUNTIFS({task},"<>",{status},"Blocked")'),
        ("Due This Week", f'=COUNTIFS({task},"<>",{plan_end},">="&TODAY(),{plan_end},"<="&TODAY()+7,{status},"<>Done",{status},"<>Cancelled")'),
    ]
    for idx, (label, formula) in enumerate(kpis, 1):
        col = 1 + (idx - 1) * 2
        ws.cell(3, col, label)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        ws.cell(4, col, formula)
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        set_header(ws.cell(3, col), "1F4E78")
        ws.cell(4, col).font = Font(name="Aptos", size=16, bold=True)
        ws.cell(4, col).alignment = Alignment(horizontal="center", vertical="center")
    style_block(ws, 3, 4, 1, 12)

    set_title(ws, 6, "Priority Queue", 9)
    queue_headers = ["Priority", "Project / Collection", "Item / Design No.", "Task", "Qty", "Status", "Owner", "Plan End", "_idx"]
    for col, header in enumerate(queue_headers, 1):
        ws.cell(7, col, header)
        set_header(ws.cell(7, col), "1F4E78")
    for row in range(8, 28):
        ws.cell(row, 9, queue_index_formula(row))
        idx_cell = f"I{row}"
        mappings = [
            (1, "Priority"),
            (2, "Project / Collection"),
            (3, "Item / Design No."),
            (4, "Task"),
            (5, "Qty"),
            (6, "Status"),
            (7, "Owner"),
            (8, "Plan End"),
        ]
        for col, source_col in mappings:
            ws.cell(row, col, queue_value_formula(idx_cell, source_col))
        ws.cell(row, 8).number_format = "[$-en-US]d mmm yy"
    ws.column_dimensions["I"].hidden = True
    style_block(ws, 6, 27, 1, 9)

    set_title(ws, 29, "Owner Workload", 5)
    owner_headers = ["Owner", "Active Tasks", "Active Items", "Urgent", "Blocked"]
    for col, header in enumerate(owner_headers, 1):
        ws.cell(30, col, header)
        set_header(ws.cell(30, col), "1F4E78")
    for row, owner_name in enumerate(OWNERS, 31):
        owner_cell = f"A{row}"
        ws.cell(row, 1, owner_name)
        ws.cell(row, 2, f'=COUNTIFS({task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled")')
        ws.cell(row, 3, qty_active_formula(owner_cell))
        ws.cell(row, 4, f'=COUNTIFS({task},"<>",{owner},{owner_cell},{priority},"Urgent",{status},"<>Done",{status},"<>Cancelled")')
        ws.cell(row, 5, f'=COUNTIFS({task},"<>",{owner},{owner_cell},{status},"Blocked")')
    style_block(ws, 29, 39, 1, 5)

    set_title(ws, 41, "Batch Summary", 7)
    batch_headers = ["Batch ID", "Project / Collection", "Total Items", "Done Items", "Blocked Items", "Progress %", "Overall Status"]
    for col, header in enumerate(batch_headers, 1):
        ws.cell(42, col, header)
        set_header(ws.cell(42, col), "1F4E78")
    for row in range(43, 63):
        batch_cell = f"A{row}"
        total_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>")+COUNTIFS({batch},{batch_cell},{task},"<>",{qty},"")'
        done_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>",{status},"Done")+COUNTIFS({batch},{batch_cell},{task},"<>",{status},"Done",{qty},"")'
        blocked_items = f'SUMIFS({qty},{batch},{batch_cell},{task},"<>",{status},"Blocked")+COUNTIFS({batch},{batch_cell},{task},"<>",{status},"Blocked",{qty},"")'
        ws.cell(row, 1, batch_id_formula(row, 43))
        ws.cell(row, 2, f'=IF({batch_cell}="","",INDEX({project},MATCH({batch_cell},{batch},0)))')
        ws.cell(row, 3, f'=IF({batch_cell}="","",{total_items})')
        ws.cell(row, 4, f'=IF({batch_cell}="","",{done_items})')
        ws.cell(row, 5, f'=IF({batch_cell}="","",{blocked_items})')
        ws.cell(row, 6, f'=IFERROR(({done_items})/({total_items}),"")')
        ws.cell(row, 7, (
            f'=IF({batch_cell}="","",IF(({done_items})=({total_items}),"Done",'
            f'IF(({blocked_items})>0,"Blocked",'
            f'IF(COUNTIFS({batch},{batch_cell},{status},"In Progress")+COUNTIFS({batch},{batch_cell},{status},"Checking")>0,"In Progress",'
            f'IF(COUNTIFS({batch},{batch_cell},{status},"Waiting")=COUNTIFS({batch},{batch_cell},{task},"<>"),"Waiting","Mixed")))))'
        ))
        ws.cell(row, 6).number_format = "0%"
    style_block(ws, 41, 62, 1, 7)

    widths = {
        "A": 18, "B": 24, "C": 20, "D": 32, "E": 12, "F": 16,
        "G": 18, "H": 14, "I": 8, "J": 14, "K": 14, "L": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"


def main():
    wb = load_workbook(SOURCE)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    rebuild_dashboard(wb["Dashboard"])
    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    ws = check["Dashboard"]
    errors = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for token in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
                    if token in cell.value:
                        errors.append((cell.coordinate, cell.value))
    if errors:
        raise RuntimeError(errors[:10])
    print(OUTPUT)


if __name__ == "__main__":
    main()
