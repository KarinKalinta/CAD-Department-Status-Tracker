from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties


SOURCE = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_FINAL_OWNER_FIXED_source.xlsx")
OUTPUT = Path("CAD_Status_Tracker_Batch_Aware_MANAGER_VIEW_OWNER_WORKLOAD_FIXED.xlsx")


def main():
    wb = load_workbook(SOURCE)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    ws = wb["Dashboard"]

    owner_title = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Owner Workload":
                owner_title = cell
                break
        if owner_title:
            break

    if owner_title is None:
        raise RuntimeError("Owner Workload section not found")

    first_owner_row = owner_title.row + 2
    last_owner_row = first_owner_row + 8

    task = "CADTracker[Task]"
    owner = "CADTracker[Owner]"
    status = "CADTracker[Status]"
    qty = "CADTracker[Qty]"
    plan_end = "CADTracker[Plan End]"

    for row in range(first_owner_row, last_owner_row + 1):
        owner_cell = f"$A{row}"

        # Active Tasks = count active rows for this owner.
        ws.cell(row, 2).value = (
            f'=COUNTIFS({task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled")'
        )

        # Active Items = sum Qty for active rows, treating blank Qty as 1.
        ws.cell(row, 3).value = (
            f'=SUMIFS({qty},{task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled")'
            f'+COUNTIFS({task},"<>",{owner},{owner_cell},{status},"<>Done",{status},"<>Cancelled",{qty},"")'
        )

        ws.cell(row, 4).value = (
            f'=COUNTIFS({task},"<>",{owner},{owner_cell},{plan_end},"<"&TODAY(),{status},"<>Done",{status},"<>Cancelled")'
        )
        ws.cell(row, 5).value = f'=COUNTIFS({task},"<>",{owner},{owner_cell},{status},"Done")'

    wb.save(OUTPUT)

    check = load_workbook(OUTPUT, data_only=False)
    errors = []
    for row in check["Dashboard"].iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                for token in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!"):
                    if token in cell.value:
                        errors.append((cell.coordinate, cell.value))
    if errors:
        raise RuntimeError(errors[:20])

    print(OUTPUT)


if __name__ == "__main__":
    main()
