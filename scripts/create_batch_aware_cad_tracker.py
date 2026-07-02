from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties


OUTPUT_FILE = Path("CAD_Status_Tracker_Batch_Aware_OWNERS_UPDATED.xlsx")
MAX_ROWS = 201

TRACKER_HEADERS = [
    "Tracker ID",
    "Batch ID",
    "WE",
    "Project / Collection",
    "Item / Design No.",
    "Request Type",
    "Deliverable Type",
    "Task",
    "Qty",
    "Status",
    "Owner",
    "Priority",
    "Start Date",
    "Plan End",
]

LISTS = {
    "Status": ["Waiting", "In Progress", "Checking", "Done", "Blocked", "Cancelled"],
    "Request Type": ["Initial Request", "New Request", "Revision", "Follow-up", "Internal Fix"],
    "Deliverable Type": ["CAD File", "Render", "Sample", "Existing Sketch", "Photo", "File Check"],
    "Priority": ["Urgent", "High", "Normal", "Low"],
    "Owner": ["CAD1", "CAD2", "CAD3", "CAD4", "CAD5", "CAD6", "Outsource CAD", "Sdesign", "Design 2"],
}

EXAMPLES = [
    ["", "BATCH-001", "2026-W27", "C0634", "Art 01-20", "Initial Request", "CAD File", "Pendant CAD", 20, "In Progress", "CAD1", "Urgent", "2026-06-29", "2026-07-05"],
    ["", "BATCH-002", "2026-W27", "C0553-F", "Art 01", "Revision", "CAD File", "Pendant CAD revision", 1, "Done", "CAD2", "High", "2026-06-29", "2026-07-01"],
    ["", "BATCH-002", "2026-W27", "C0553-F", "Art 02", "Revision", "CAD File", "Pendant CAD revision", 1, "In Progress", "CAD3", "High", "2026-06-29", "2026-07-03"],
    ["", "BATCH-002", "2026-W27", "C0553-F", "Art 03", "Revision", "CAD File", "Pendant CAD revision", 1, "Blocked", "Sdesign", "High", "2026-06-29", "2026-07-03"],
    ["", "", "2026-W27", "PD CAD Investigation", "Check set", "Internal Fix", "File Check", "CAD file naming audit", 10, "Checking", "Design 2", "Normal", "2026-06-30", "2026-07-04"],
]


def table_col(column_name):
    names = {
        "Tracker ID": "CADTracker_Tracker_ID",
        "Batch ID": "CADTracker_Batch_ID",
        "WE": "CADTracker_WE",
        "Project / Collection": "CADTracker_Project",
        "Item / Design No.": "CADTracker_Item",
        "Request Type": "CADTracker_Request_Type",
        "Deliverable Type": "CADTracker_Deliverable_Type",
        "Task": "CADTracker_Task",
        "Qty": "CADTracker_Qty",
        "Status": "CADTracker_Status",
        "Owner": "CADTracker_Owner",
        "Priority": "CADTracker_Priority",
        "Start Date": "CADTracker_Start_Date",
        "Plan End": "CADTracker_Plan_End",
    }
    return names[column_name]


def style_header(range_obj, fill="1F4E78"):
    range_obj.fill = PatternFill("solid", fgColor=fill)
    range_obj.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    range_obj.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_styles():
    thin = Side(style="thin", color="D9E2F3")
    return {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "dark": PatternFill("solid", fgColor="111827"),
        "blue": PatternFill("solid", fgColor="1F4E78"),
        "light": PatternFill("solid", fgColor="F8FAFC"),
    }


def setup_lists(ws):
    for col_idx, (header, values) in enumerate(LISTS.items(), 1):
        cell = ws.cell(1, col_idx, header)
        style_header(cell)
        for row_idx, value in enumerate(values, 2):
            ws.cell(row_idx, col_idx, value)
        ws.column_dimensions[get_column_letter(col_idx)].width = 24
    ws.sheet_state = "hidden"


def add_dropdown(ws, target_range, list_col):
    formula = f"='Lists hidden'!${list_col}$2:${list_col}$100"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Choose a value from the dropdown list."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(target_range)


def setup_tracker(ws, include_examples=False, table_name="CADTracker"):
    styles = create_styles()
    ws.append(TRACKER_HEADERS)
    for cell in ws[1]:
        style_header(cell)
        cell.border = styles["border"]

    widths = [14, 16, 12, 24, 20, 18, 18, 28, 10, 16, 16, 12, 13, 13]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in range(2, MAX_ROWS + 1):
        ws.cell(row, 1).value = f'=IF($H{row}<>"","CAD-"&TEXT(ROW()-1,"0000"),"")'
        ws.cell(row, 9).number_format = "0"
        for col in (13, 14):
            ws.cell(row, col).number_format = "[$-en-US]d mmm yy"
        for col in range(1, len(TRACKER_HEADERS) + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Aptos", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=col == 8)
            cell.border = styles["border"]

    if include_examples:
        for row_offset, row_values in enumerate(EXAMPLES, 2):
            for col_idx, value in enumerate(row_values, 1):
                ws.cell(row_offset, col_idx, value)
            ws.cell(row_offset, 1).value = f'=IF($H{row_offset}<>"","CAD-"&TEXT(ROW()-1,"0000"),"")'

    table = Table(displayName=table_name, ref=f"A1:N{MAX_ROWS}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"

    add_dropdown(ws, f"F2:F{MAX_ROWS}", "B")
    add_dropdown(ws, f"G2:G{MAX_ROWS}", "C")
    add_dropdown(ws, f"J2:J{MAX_ROWS}", "A")
    add_dropdown(ws, f"K2:K{MAX_ROWS}", "E")
    add_dropdown(ws, f"L2:L{MAX_ROWS}", "D")

    add_color_rules(ws)
    apply_existing_color_fills(ws)


def setup_tracker_names(wb):
    for column_index, header in enumerate(TRACKER_HEADERS, 1):
        column_letter = get_column_letter(column_index)
        defined_name = DefinedName(
            table_col(header),
            attr_text=f"'CAD Tracker'!${column_letter}$2:${column_letter}${MAX_ROWS}",
        )
        wb.defined_names.add(defined_name)


def color_rules():
    return {
        "J": {
            "Waiting": ("FEF3C7", "92400E"),
            "In Progress": ("DBEAFE", "1E3A8A"),
            "Checking": ("EDE9FE", "5B21B6"),
            "Done": ("DCFCE7", "166534"),
            "Blocked": ("FEE2E2", "991B1B"),
            "Cancelled": ("E5E7EB", "374151"),
        },
        "K": {
            "CAD1": ("FED7AA", "9A3412"),
            "CAD2": ("DBEAFE", "1E3A8A"),
            "CAD3": ("DCFCE7", "166534"),
            "CAD4": ("EDE9FE", "5B21B6"),
            "CAD5": ("FEF3C7", "92400E"),
            "CAD6": ("E0F2FE", "075985"),
            "Outsource CAD": ("E5E7EB", "374151"),
            "Sdesign": ("FCE7F3", "9D174D"),
            "Design 2": ("F3E8FF", "6B21A8"),
        },
        "L": {
            "Urgent": ("FCA5A5", "7F1D1D"),
            "High": ("FFEDD5", "9A3412"),
            "Normal": ("E5E7EB", "374151"),
            "Low": ("F8FAFC", "64748B"),
        },
    }


def add_color_rules(ws):
    rules = color_rules()
    for column, values in rules.items():
        target = f"{column}2:{column}{MAX_ROWS}"
        for value, (fill, font_color) in values.items():
            ws.conditional_formatting.add(
                target,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{value}"'],
                    fill=PatternFill(fill_type="solid", start_color=fill, end_color=fill),
                    font=Font(name="Aptos", size=10, color=font_color),
                ),
            )


def apply_existing_color_fills(ws):
    rules = color_rules()
    for column, values in rules.items():
        for row in range(2, MAX_ROWS + 1):
            cell = ws[f"{column}{row}"]
            if cell.value in values:
                fill, font_color = values[cell.value]
                cell.fill = PatternFill(fill_type="solid", start_color=fill, end_color=fill)
                cell.font = Font(name="Aptos", size=10, color=font_color)


def setup_dashboard(ws):
    styles = create_styles()
    task = table_col("Task")
    qty = table_col("Qty")
    status = table_col("Status")
    plan_end = table_col("Plan End")
    batch_id = table_col("Batch ID")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "CAD Batch Status Dashboard"
    ws["A1"].font = Font(name="Aptos", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = styles["dark"]
    ws.merge_cells("A1:N1")

    kpis = [
        ("Total Tasks", f'=COUNTIFS({task},"<>")'),
        ("Total Items", f'=SUMIFS({qty},{task},"<>")+COUNTIFS({task},"<>",{qty},"")'),
        ("Active Tasks", f'=COUNTIFS({task},"<>",{status},"<>Done",{status},"<>Cancelled")'),
        ("Waiting", f'=COUNTIFS({task},"<>",{status},"Waiting")'),
        ("In Progress", f'=COUNTIFS({task},"<>",{status},"In Progress")'),
        ("Checking", f'=COUNTIFS({task},"<>",{status},"Checking")'),
        ("Blocked", f'=COUNTIFS({task},"<>",{status},"Blocked")'),
        ("Done", f'=COUNTIFS({task},"<>",{status},"Done")'),
        ("Overdue", f'=COUNTIFS({task},"<>",{plan_end},"<"&TODAY(),{status},"<>Done",{status},"<>Cancelled")'),
        ("Due This Week", f'=COUNTIFS({task},"<>",{plan_end},">="&TODAY(),{plan_end},"<="&TODAY()+7,{status},"<>Done",{status},"<>Cancelled")'),
        ("Batch Count", f'=IFERROR(COUNTA(UNIQUE(FILTER({batch_id},{batch_id}<>""))),0)'),
    ]

    for idx, (label, formula) in enumerate(kpis):
        col = 1 + (idx % 6) * 2
        row = 3 + (idx // 6) * 3
        ws.cell(row, col, label)
        ws.cell(row + 1, col, formula)
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        style_header(ws.cell(row, col), "1F4E78")
        ws.cell(row + 1, col).font = Font(name="Aptos", size=16, bold=True)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")

    sections = [
        ("Status Summary", 9, 1, ["Status", "Task Count", "Item Count"], LISTS["Status"]),
        ("Request Type Split", 9, 5, ["Request Type", "Task Count", "Item Count"], LISTS["Request Type"]),
        ("Deliverable Type Split", 9, 9, ["Deliverable Type", "Task Count", "Item Count"], LISTS["Deliverable Type"]),
        ("Owner Workload", 20, 1, ["Owner", "Active Tasks", "Active Items", "Overdue", "Done"], LISTS["Owner"]),
    ]
    for title, start_row, start_col, headers, labels in sections:
        write_summary_section(ws, title, start_row, start_col, headers, labels)

    write_batch_summary(ws, 20, 8)

    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 17


def add_dashboard_charts(ws):
    ws["A36"] = "Visual Status View"
    ws["A36"].font = Font(name="Aptos", size=14, bold=True, color="FFFFFF")
    ws["A36"].fill = PatternFill("solid", fgColor="111827")
    ws.merge_cells("A36:N36")

    status_donut = DoughnutChart()
    status_donut.title = "Task Status Split"
    status_donut.holeSize = 55
    status_donut.add_data(Reference(ws, min_col=2, min_row=10, max_row=16), titles_from_data=True)
    status_donut.set_categories(Reference(ws, min_col=1, min_row=11, max_row=16))
    status_donut.height = 7
    status_donut.width = 9
    ws.add_chart(status_donut, "A38")

    status_items = BarChart()
    status_items.type = "bar"
    status_items.style = 10
    status_items.title = "Items by Status"
    status_items.y_axis.title = "Status"
    status_items.x_axis.title = "Items"
    status_items.add_data(Reference(ws, min_col=3, min_row=10, max_row=16), titles_from_data=True)
    status_items.set_categories(Reference(ws, min_col=1, min_row=11, max_row=16))
    status_items.height = 7
    status_items.width = 10
    ws.add_chart(status_items, "H38")

    owner_workload = BarChart()
    owner_workload.type = "bar"
    owner_workload.style = 11
    owner_workload.title = "Active Items by Owner"
    owner_workload.y_axis.title = "Owner"
    owner_workload.x_axis.title = "Active Items"
    owner_workload.add_data(Reference(ws, min_col=3, min_row=21, max_row=28), titles_from_data=True)
    owner_workload.set_categories(Reference(ws, min_col=1, min_row=22, max_row=28))
    owner_workload.height = 7
    owner_workload.width = 9
    ws.add_chart(owner_workload, "A55")

    deliverable_split = BarChart()
    deliverable_split.type = "bar"
    deliverable_split.style = 12
    deliverable_split.title = "Items by Deliverable"
    deliverable_split.y_axis.title = "Deliverable"
    deliverable_split.x_axis.title = "Items"
    deliverable_split.add_data(Reference(ws, min_col=11, min_row=10, max_row=16), titles_from_data=True)
    deliverable_split.set_categories(Reference(ws, min_col=9, min_row=11, max_row=16))
    deliverable_split.height = 7
    deliverable_split.width = 10
    ws.add_chart(deliverable_split, "H55")


def write_summary_section(ws, title, start_row, start_col, headers, labels):
    task = table_col("Task")
    qty = table_col("Qty")
    status = table_col("Status")
    request_type = table_col("Request Type")
    deliverable_type = table_col("Deliverable Type")
    owner = table_col("Owner")
    plan_end = table_col("Plan End")
    ws.cell(start_row, start_col, title)
    style_header(ws.cell(start_row, start_col), "111827")
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
    for idx, header in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + idx, header)
        style_header(cell)
    for offset, label in enumerate(labels, start_row + 2):
        ws.cell(offset, start_col, label)
        first_header = headers[0]
        label_cell = ws.cell(offset, start_col).coordinate
        if first_header == "Status":
            ws.cell(offset, start_col + 1, f'=COUNTIFS({task},"<>",{status},{label_cell})')
            ws.cell(offset, start_col + 2, f'=SUMIFS({qty},{task},"<>",{status},{label_cell})+COUNTIFS({task},"<>",{status},{label_cell},{qty},"")')
        elif first_header == "Request Type":
            ws.cell(offset, start_col + 1, f'=COUNTIFS({task},"<>",{request_type},{label_cell})')
            ws.cell(offset, start_col + 2, f'=SUMIFS({qty},{task},"<>",{request_type},{label_cell})+COUNTIFS({task},"<>",{request_type},{label_cell},{qty},"")')
        elif first_header == "Deliverable Type":
            ws.cell(offset, start_col + 1, f'=COUNTIFS({task},"<>",{deliverable_type},{label_cell})')
            ws.cell(offset, start_col + 2, f'=SUMIFS({qty},{task},"<>",{deliverable_type},{label_cell})+COUNTIFS({task},"<>",{deliverable_type},{label_cell},{qty},"")')
        elif first_header == "Owner":
            ws.cell(offset, start_col + 1, f'=COUNTIFS({task},"<>",{owner},{label_cell},{status},"<>Done",{status},"<>Cancelled")')
            ws.cell(offset, start_col + 2, f'=SUMIFS({qty},{task},"<>",{owner},{label_cell},{status},"<>Done",{status},"<>Cancelled")+COUNTIFS({task},"<>",{owner},{label_cell},{status},"<>Done",{status},"<>Cancelled",{qty},"")')
            ws.cell(offset, start_col + 3, f'=COUNTIFS({task},"<>",{owner},{label_cell},{plan_end},"<"&TODAY(),{status},"<>Done",{status},"<>Cancelled")')
            ws.cell(offset, start_col + 4, f'=COUNTIFS({task},"<>",{owner},{label_cell},{status},"Done")')


def write_batch_summary(ws, start_row, start_col):
    task = table_col("Task")
    qty = table_col("Qty")
    status = table_col("Status")
    batch_id = table_col("Batch ID")
    project = table_col("Project / Collection")
    plan_end = table_col("Plan End")
    headers = ["Batch ID", "Project / Collection", "Total Rows", "Total Items", "Done Items", "Blocked Items", "% Complete", "Latest Plan End", "Overall Status"]
    ws.cell(start_row, start_col, "Batch Summary")
    style_header(ws.cell(start_row, start_col), "111827")
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
    for idx, header in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + idx, header)
        style_header(cell)

    first_data = start_row + 2
    for row in range(first_data, first_data + 30):
        batch_cell = ws.cell(row, start_col).coordinate
        previous_batch_range = f"${get_column_letter(start_col)}${first_data - 1}:{get_column_letter(start_col)}{row - 1}"
        ws.cell(row, start_col, f'=IFERROR(INDEX({batch_id},MATCH(0,COUNTIF({previous_batch_range},{batch_id})+({batch_id}=""),0)),"")')
        ws.cell(row, start_col + 1, f'=IF({batch_cell}="","",INDEX({project},MATCH({batch_cell},{batch_id},0)))')
        ws.cell(row, start_col + 2, f'=IF({batch_cell}="","",COUNTIFS({batch_id},{batch_cell},{task},"<>"))')
        total_items_formula = f'SUMIFS({qty},{batch_id},{batch_cell},{task},"<>")+COUNTIFS({batch_id},{batch_cell},{task},"<>",{qty},"")'
        done_items_formula = f'SUMIFS({qty},{batch_id},{batch_cell},{task},"<>",{status},"Done")+COUNTIFS({batch_id},{batch_cell},{task},"<>",{status},"Done",{qty},"")'
        blocked_items_formula = f'SUMIFS({qty},{batch_id},{batch_cell},{task},"<>",{status},"Blocked")+COUNTIFS({batch_id},{batch_cell},{task},"<>",{status},"Blocked",{qty},"")'
        ws.cell(row, start_col + 3, f'=IF({batch_cell}="","",{total_items_formula})')
        ws.cell(row, start_col + 4, f'=IF({batch_cell}="","",{done_items_formula})')
        ws.cell(row, start_col + 5, f'=IF({batch_cell}="","",{blocked_items_formula})')
        ws.cell(
            row,
            start_col + 6,
            (
                f'=IFERROR('
                f'({done_items_formula})/({total_items_formula}),'
                f'"")'
            ),
        )
        ws.cell(row, start_col + 7, f'=IF({batch_cell}="","",MAXIFS({plan_end},{batch_id},{batch_cell},{task},"<>"))')
        ws.cell(row, start_col + 8, (
            f'=IF({batch_cell}="","",IF('
            f'({done_items_formula})=({total_items_formula}),"Done",'
            f'IF(({blocked_items_formula})>0,"Blocked",'
            f'IF(COUNTIFS({batch_id},{batch_cell},{status},"In Progress")+COUNTIFS({batch_id},{batch_cell},{status},"Checking")>0,"In Progress",'
            f'IF(COUNTIFS({batch_id},{batch_cell},{status},"Waiting")=COUNTIFS({batch_id},{batch_cell},{task},"<>"),"Waiting","Mixed")))))'
        ))
        ws.cell(row, start_col + 6).number_format = "0%"
        ws.cell(row, start_col + 7).number_format = "[$-en-US]d mmm yy"


def setup_readme(ws):
    ws["A1"] = "CAD Batch-Aware Tracker README"
    ws["A1"].font = Font(name="Aptos", size=16, bold=True)
    notes = [
        "One row = one trackable work unit.",
        "Use one row with Qty > 1 if the whole batch moves together.",
        "Use multiple rows with the same Batch ID if items move separately.",
        "Task must be filled for dashboard counting.",
        "Qty controls item-based progress. If Qty is blank, dashboard formulas treat it as 1.",
        "Tracker ID is auto-generated from the row when Task is filled.",
        "Batch ID can be blank for single non-batch tasks.",
    ]
    for idx, note in enumerate(notes, 3):
        ws.cell(idx, 1, note)
    ws.column_dimensions["A"].width = 100


def create_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    tracker = wb.create_sheet("CAD Tracker")
    dashboard = wb.create_sheet("Dashboard", 0)
    examples = wb.create_sheet("CAD Examples")
    lists = wb.create_sheet("Lists hidden")
    readme = wb.create_sheet("README")

    setup_lists(lists)
    setup_tracker(tracker, include_examples=True, table_name="CADTracker")
    setup_tracker(examples, include_examples=True, table_name="CADExamples")
    setup_tracker_names(wb)
    setup_dashboard(dashboard)
    setup_readme(readme)

    dashboard.sheet_view.topLeftCell = "A1"
    wb.save(OUTPUT_FILE)


if __name__ == "__main__":
    create_workbook()
